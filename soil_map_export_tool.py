#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import traceback
import zipfile
from collections import OrderedDict, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from xml.etree import ElementTree as ET

# =============================================================================
# 【用户路径配置区】—— 部署时按实际情况修改以下常量，程序运行时以此为默认值
# =============================================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# 各县图框模板（.mxd）存放的父目录；每个县的 mxd 应位于其下的子目录中
CFG_FRAMES_BASE_DIR   = PROJECT_ROOT / "图框"

# 各县土壤类型 .gdb 数据库存放的父目录
CFG_SOIL_GDB_BASE_DIR = PROJECT_ROOT / "土壤类型"

# 配色表（全局唯一，固定路径）
CFG_COLOR_TABLE       = PROJECT_ROOT / "a出图程序" / "土壤类型配色表.csv"

# 平差面积统计 Excel 存放目录（各县 xlsx 均置于此处）
CFG_STATS_BASE_DIR    = PROJECT_ROOT / "三普土壤类型统计表" / "xx县土壤类型统计表.xlsx"

# 成果输出目录（布局工程、JPG、.lyrx、配色对照表 Excel 均输出至此）
CFG_OUTPUT_BASE_DIR   = PROJECT_ROOT / "出图工程" / "xx县出图工程文件"

# =============================================================================
# 全局常量（一般无需修改）
# =============================================================================

# 统计 Excel 中面积列的字段名；值的原始单位：平方米
STATS_AREA_FIELD = "面积"
SQM_TO_MU = 1.0 / 666.6667  # 平方米 → 亩 换算系数

LAYER_NAME_SUFFIX = "土壤类型配色表"  # 动态图层名后缀，前缀为县名
FULL_DOMAIN_SUFFIX = "土壤类型分布图"  # 全域图名后缀

DEFAULT_SOIL_FIELD = "TL"
DEFAULT_SUBCLASS_FIELD = "YL"
DEFAULT_GENUS_FIELD = "TS"
DEFAULT_EXCLUDES = "建设用地,河流水面"
DEFAULT_DPI = 600
DEFAULT_AREA_UNIT = "亩"                    # 面积统计单位："亩" 或 "万亩"
DEFAULT_COLOR_DIRECTION = "dark_to_light"   # 色带方向："dark_to_light"（深→浅）或 "light_to_dark"（浅→深）

AREA_UNIT_LABELS = {"亩": "面积/亩", "万亩": "面积/万亩"}
AREA_UNIT_FACTORS = {"亩": 1.0, "万亩": 1.0 / 10000.0}
COLOR_DIRECTION_CHOICES = {"dark_to_light": "深→浅", "light_to_dark": "浅→深"}


# =============================================================================
# 第一部分：环境初始化（延迟加载，仅在实际出图时才导入 arcpy / openpyxl）
# =============================================================================


def _install_package(package: str) -> bool:
    """通过 pip 安装指定包。"""
    print(f"  正在安装 {package} ...")
    try:
        r = subprocess.run(
            [sys.executable, "-m", "pip", "install", package],
            capture_output=True, text=True, encoding="gbk", errors="replace", timeout=120,
        )
        if r.returncode == 0:
            print(f"  {package} 安装成功。")
            return True
    except FileNotFoundError:
        pass
    except Exception as e:
        print(f"  pip 安装失败：{e}")

    print(f"  尝试使用 ensurepip ...")
    try:
        subprocess.run(
            [sys.executable, "-m", "ensurepip", "--upgrade"],
            capture_output=True, timeout=60,
        )
        r = subprocess.run(
            [sys.executable, "-m", "pip", "install", package],
            capture_output=True, text=True, encoding="gbk", errors="replace", timeout=120,
        )
        if r.returncode == 0:
            print(f"  {package} 安装成功（通过 ensurepip）。")
            return True
    except Exception:
        pass

    print(f"  自动安装失败，请手动执行：\n  \"{sys.executable}\" -m pip install {package}")
    return False


def _ensure_openpyxl() -> None:
    """确保 openpyxl 可用，缺失时自动安装。"""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("openpyxl 未安装，正在自动安装...")
        if not _install_package("openpyxl"):
            raise RuntimeError(
                f"openpyxl 安装失败，请手动执行：\n\"{sys.executable}\" -m pip install openpyxl"
            )


# =============================================================================
# 第二部分：配色表读取与颜色计算
# =============================================================================

def read_color_table(path: Path) -> Dict[str, List[Tuple[int, int, int]]]:
    """
    读取 CSV 配色表（自动检测 GBK/UTF-8），返回 {土类: [(R,G,B),...]} 字典。
    """
    required = {"土类", "R", "G", "B"}

    for enc in ("gbk", "utf-8-sig", "utf-8"):
        try:
            palettes: Dict[str, List[Tuple[int, int, int]]] = defaultdict(list)
            with path.open("r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames or not required.issubset(set(reader.fieldnames)):
                    raise ValueError(f"配色表须含字段：{required}；当前表头：{reader.fieldnames}")
                rows = sorted(
                    reader,
                    key=lambda r: (str(r.get("土类", "")), int(float(r.get("色标序号") or 0))),
                )
                for row in rows:
                    sc = str(row["土类"]).strip()
                    if sc:
                        palettes[sc].append((
                            int(float(row["R"])),
                            int(float(row["G"])),
                            int(float(row["B"])),
                        ))
            if not palettes:
                raise RuntimeError(
                    f"配色表解析成功（编码 {enc}）但未读取到任何土类配色行，"
                    f"请检查文件内容：{path}"
                )
            return dict(palettes)
        except RuntimeError:
            raise
        except Exception:
            continue

    raise RuntimeError(f"读取配色表失败（已尝试 gbk/utf-8-sig/utf-8）：{path}")


def _lerp(
        a: Tuple[int, int, int],
        b: Tuple[int, int, int],
        i: int,
        n: int,
) -> Tuple[int, int, int]:
    """在两个 RGB 端点之间做线性插值。"""
    if n <= 1:
        return a
    t = i / float(n - 1)
    return (
        int(round(a[0] + (b[0] - a[0]) * t)),
        int(round(a[1] + (b[1] - a[1]) * t)),
        int(round(a[2] + (b[2] - a[2]) * t)),
    )


def _shade_color(rgb: Tuple[int, int, int], i: int, n: int) -> Tuple[int, int, int]:
    """当配色表只有 1 个锚点色时，生成一组同色系明暗变化色。"""
    if n <= 1:
        return rgb
    factor = 1.18 - 0.36 * (i / float(n - 1))
    return tuple(max(0, min(255, int(round(c * factor)))) for c in rgb)  # type: ignore[return-value]


def _palette_pick_order(count: int, palette_len: int) -> List[int]:
    """
    返回按用户规则选择的 0-based 色号顺序。
    """
    if count <= 0 or palette_len <= 0 or count > palette_len:
        return []
    start_no = 4 if count <= 4 else count
    if start_no > palette_len:
        return []
    return [color_no - 1 for color_no in range(start_no, start_no - count, -1)]


def _select_palette_colors(
        palette: List[Tuple[int, int, int]],
        count: int,
        color_direction: str = DEFAULT_COLOR_DIRECTION,
) -> List[Tuple[int, int, int]]:
    """优先按指定色号取色；已有配色不够时，再插值生成足量颜色。"""
    if count <= 0:
        return []
    if not palette:
        return [(210, 210, 210)] * count

    order = _palette_pick_order(count, len(palette))
    if order:
        colors = [palette[i] for i in order]
    elif len(palette) >= 2:
        colors = [_lerp(palette[-1], palette[0], i, count) for i in range(count)]
    else:
        colors = [_shade_color(palette[0], i, count) for i in range(count)]

    if color_direction == "light_to_dark":
        colors = colors[::-1]
    return colors


def build_value_colors(
        soil_class: str,
        stats_rows: List[Dict],
        palettes: Dict[str, List[Tuple[int, int, int]]],
        use_subclass_mode: bool,
        color_direction: str = DEFAULT_COLOR_DIRECTION,
) -> Dict[str, Tuple[int, int, int]]:
    """为指定土类的亚类或土属分配颜色。"""
    palette = palettes.get(soil_class) or []
    key = "亚类" if use_subclass_mode else "土属"
    seen: set = set()
    names: List[str] = []
    for row in stats_rows:
        v = str(row[key]).strip()
        if v and v not in seen:
            names.append(v)
            seen.add(v)
    n = len(names)
    if n == 0:
        return {}
    if not palette:
        print(
            f"警告：土类「{soil_class}」在配色表中无任何颜色记录，所有{key}值将使用默认灰色。",
            file=sys.stderr,
        )
    elif n > len(palette):
        print(
            f"提示：土类「{soil_class}」存在 {n} 个{key}值，配色表仅有 {len(palette)} 个颜色，"
            "将使用插值补足。",
            file=sys.stderr,
        )
    colors = _select_palette_colors(palette, n, color_direction)
    return dict(zip(names, colors))


def build_full_subclass_colors(
        soil_classes: List[str],
        stats_by_soil: Dict[str, List[Dict]],
        palettes: Dict[str, List[Tuple[int, int, int]]],
        color_direction: str = DEFAULT_COLOR_DIRECTION,
) -> Dict[str, Tuple[int, int, int]]:
    """构建全域图的亚类颜色映射（各土类按自身调色盘分段插值后合并）。"""
    merged: Dict[str, Tuple[int, int, int]] = {}
    for sc in soil_classes:
        merged.update(build_value_colors(sc, stats_by_soil.get(sc, []), palettes, True, color_direction))
    return merged


# =============================================================================
# 第三部分：统计表读取（含平方米 → 亩 换算）
# =============================================================================

def read_stats_table(path: Path) -> List[Dict]:
    """读取平差面积统计 Excel，面积字段自动换算为亩。"""
    try:
        import openpyxl  # type: ignore
        return _read_with_openpyxl(path, openpyxl)
    except ImportError:
        return _read_without_openpyxl(path)


def _read_with_openpyxl(path: Path, openpyxl) -> List[Dict]:
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    return _parse_rows(headers, (list(r) for r in ws.iter_rows(min_row=2, values_only=True)))


def _read_without_openpyxl(path: Path) -> List[Dict]:
    raw = _parse_xlsx_xml(path)
    if not raw:
        raise RuntimeError(f"统计表为空：{path}")
    return _parse_rows(
        [str(v).strip() if v is not None else "" for v in raw[0]],
        raw[1:],
    )


def _parse_rows(headers: List[str], data_rows: Iterable) -> List[Dict]:
    """将统计表原始行解析为字典列表，面积换算为亩。"""
    required = ["土类", "亚类", "土属", STATS_AREA_FIELD]
    missing = [f for f in required if f not in headers]
    if missing:
        raise RuntimeError(
            f"统计表缺少字段：{', '.join(missing)}；当前表头：{', '.join(headers)}"
        )
    idx = {f: headers.index(f) for f in required}
    result: List[Dict] = []
    last_soil = last_sub = ""
    for row in data_rows:
        row = list(row)
        pad = row + [""] * max(0, len(headers) - len(row))
        soil = pad[idx["土类"]]
        subclass = pad[idx["亚类"]]
        genus = pad[idx["土属"]]
        area_sqm = pad[idx[STATS_AREA_FIELD]]
        if soil not in (None, ""):
            last_soil = str(soil).strip()
        if subclass not in (None, ""):
            last_sub = str(subclass).strip()
        if not last_soil:
            continue
        result.append({
            "土类": last_soil,
            "亚类": last_sub,
            "土属": "" if genus in (None, "") else str(genus).strip(),
            "面积_亩": float(area_sqm or 0) * SQM_TO_MU,
        })
    return result


def _parse_xlsx_xml(path: Path) -> List[List]:
    """零依赖解析 .xlsx（Office Open XML）。"""
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    with zipfile.ZipFile(str(path)) as zf:
        shared: List[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            for si in ET.fromstring(zf.read("xl/sharedStrings.xml")).findall("main:si", ns):
                shared.append("".join(t.text or "" for t in si.findall(".//main:t", ns)))

        wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
        first = wb_root.find("main:sheets/main:sheet", ns)
        if first is None:
            return []
        rid = first.attrib.get(
            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        )
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        target = next(
            (r.attrib.get("Target") for r in rels.findall("pkgrel:Relationship", ns)
             if r.attrib.get("Id") == rid),
            None,
        )
        if not target:
            return []
        sheet_root = ET.fromstring(zf.read(f"xl/{target.lstrip('/')}"))

    result: List[List] = []
    for row_el in sheet_root.findall(".//main:sheetData/main:row", ns):
        col_map: Dict[int, object] = {}
        for c in row_el.findall("main:c", ns):
            ci = _col_idx(c.attrib.get("r", ""))
            ct = c.attrib.get("t")
            vn = c.find("main:v", ns)
            isn = c.find("main:is/main:t", ns)
            if ct == "s" and vn is not None:
                val: object = shared[int(vn.text or 0)]
            elif ct == "inlineStr" and isn is not None:
                val = isn.text or ""
            elif vn is not None:
                try:
                    val = float(vn.text or "")
                except ValueError:
                    val = vn.text or ""
            else:
                val = ""
            col_map[ci] = val
        if col_map:
            mx = max(col_map)
            result.append([col_map.get(i, "") for i in range(mx + 1)])
    return result


def _col_idx(ref: str) -> int:
    letters = "".join(c for c in ref if c.isalpha())
    if not letters:
        return 0
    v = 0
    for c in letters:
        v = v * 26 + (ord(c.upper()) - ord("A") + 1)
    return max(v - 1, 0)


# =============================================================================
# 第四部分：数据聚合工具函数
# =============================================================================

def split_names(text: str) -> List[str]:
    return [p.strip() for p in text.replace("，", ",").split(",") if p.strip()]


def list_field_values(arcpy, fc: str, field: str, excludes: List[str]) -> List[str]:
    excl = set(excludes)
    vals: set = set()
    with arcpy.da.SearchCursor(fc, [field]) as cur:
        for (v,) in cur:
            t = "" if v is None else str(v).strip()
            if t and t not in excl:
                vals.add(t)
    return sorted(vals)


def _clean_value(value) -> str:
    return "" if value is None else str(value).strip()


def list_layer_type_sets(
        arcpy,
        fc: str,
        soil_field: str,
        subclass_field: str,
        genus_field: str,
        excludes: List[str],
) -> Dict[str, set]:
    """读取图层中的土类、土类+亚类、土类+亚类+土属唯一值集合。"""
    fields = [soil_field, subclass_field] + ([genus_field] if genus_field else [])
    excl = set(excludes)
    soils: set = set()
    subclasses: set = set()
    genera: set = set()

    with arcpy.da.SearchCursor(fc, fields) as cur:
        for row in cur:
            soil = _clean_value(row[0])
            subclass = _clean_value(row[1]) if len(row) > 1 else ""
            genus = _clean_value(row[2]) if genus_field and len(row) > 2 else ""
            if not soil or soil in excl:
                continue
            soils.add(soil)
            if subclass:
                subclasses.add((soil, subclass))
            if genus_field and subclass and genus:
                genera.add((soil, subclass, genus))

    return {"soil": soils, "subclass": subclasses, "genus": genera}


@dataclass
class TypeMismatchReport:
    stats_only_soils: List[str]
    layer_only_soils: List[str]
    stats_only_subclasses: List[Tuple[str, str]]
    layer_only_subclasses: List[Tuple[str, str]]
    stats_only_genera: List[Tuple[str, str, str]]
    layer_only_genera: List[Tuple[str, str, str]]

    @property
    def has_mismatch(self) -> bool:
        return any((
            self.stats_only_soils,
            self.layer_only_soils,
            self.stats_only_subclasses,
            self.layer_only_subclasses,
            self.stats_only_genera,
            self.layer_only_genera,
        ))

    def to_text(self, limit: int = 50) -> str:
        lines = ["发现统计表与图层数据的土壤类型不一致："]

        def add_values(title: str, values: List, formatter) -> None:
            if not values:
                return
            shown = values[:limit]
            text = "；".join(formatter(v) for v in shown)
            if len(values) > limit:
                text += f"；……另有 {len(values) - limit} 项"
            lines.append(f"- {title}：{text}")

        add_values("统计表有、图层无的土类", self.stats_only_soils, lambda v: v)
        add_values("图层有、统计表无的土类", self.layer_only_soils, lambda v: v)
        add_values("统计表有、图层无的亚类", self.stats_only_subclasses, lambda v: f"{v[0]} / {v[1]}")
        add_values("图层有、统计表无的亚类", self.layer_only_subclasses, lambda v: f"{v[0]} / {v[1]}")
        add_values("统计表有、图层无的土属", self.stats_only_genera, lambda v: f"{v[0]} / {v[1]} / {v[2]}")
        add_values("图层有、统计表无的土属", self.layer_only_genera, lambda v: f"{v[0]} / {v[1]} / {v[2]}")
        lines.append("选择继续将按原逻辑运行：土类取交集出图，缺少颜色映射的亚类/土属在地图中透明处理。")
        return "\n".join(lines)


def build_type_mismatch_report(
        stats_by_soil: Dict[str, List[Dict]],
        layer_sets: Dict[str, set],
        compare_genus: bool,
) -> TypeMismatchReport:
    stats_soils = {str(sc).strip() for sc in stats_by_soil if str(sc).strip()}
    stats_subclasses = {
        (str(row.get("土类", "")).strip(), str(row.get("亚类", "")).strip())
        for rows in stats_by_soil.values()
        for row in rows
        if str(row.get("土类", "")).strip() and str(row.get("亚类", "")).strip()
    }
    stats_genera = {
        (
            str(row.get("土类", "")).strip(),
            str(row.get("亚类", "")).strip(),
            str(row.get("土属", "")).strip(),
        )
        for rows in stats_by_soil.values()
        for row in rows
        if str(row.get("土类", "")).strip()
        and str(row.get("亚类", "")).strip()
        and str(row.get("土属", "")).strip()
    }

    layer_soils = set(layer_sets.get("soil", set()))
    layer_subclasses = set(layer_sets.get("subclass", set()))
    layer_genera = set(layer_sets.get("genus", set())) if compare_genus else set()

    return TypeMismatchReport(
        stats_only_soils=sorted(stats_soils - layer_soils),
        layer_only_soils=sorted(layer_soils - stats_soils),
        stats_only_subclasses=sorted(stats_subclasses - layer_subclasses),
        layer_only_subclasses=sorted(layer_subclasses - stats_subclasses),
        stats_only_genera=sorted(stats_genera - layer_genera) if compare_genus else [],
        layer_only_genera=sorted(layer_genera - stats_genera) if compare_genus else [],
    )


def confirm_type_mismatch(report: TypeMismatchReport, args) -> None:
    if not report.has_mismatch:
        return

    text = report.to_text()
    print(text, file=sys.stderr)

    if getattr(args, "continue_on_mismatch", False):
        print("已设置 --continue-on-mismatch，继续生成布局工程。", file=sys.stderr)
        return

    # GUI 模式：弹窗必须在主线程执行（出图运行于后台子线程，子线程内新建
    # tk.Tk()/messagebox 会与主线程 mainloop 死锁，导致界面一直“运行中”无结果）。
    # 这里通过 args.confirm_dialog 回调，把确认窗口调度回主线程。
    dialog = getattr(args, "confirm_dialog", None)
    if callable(dialog):
        if dialog(text + "\n\n是否继续？"):
            print("用户选择继续，按原逻辑生成布局工程。", file=sys.stderr)
            return
        raise RuntimeError("用户取消：统计表与图层数据不一致，已终止。")

    if getattr(args, "confirm_mismatch", False):
        # 命令行/主线程场景：可直接新建临时 Tk 弹窗确认。
        try:
            import tkinter as tk
            from tkinter import messagebox
            root = tk.Tk()
            root.withdraw()
            ok = messagebox.askyesno("土壤类型不一致", text + "\n\n是否继续？")
            root.destroy()
            if ok:
                print("用户选择继续，按原逻辑生成布局工程。", file=sys.stderr)
                return
        except Exception as e:
            print(f"警告：无法弹出确认窗口：{e}", file=sys.stderr)

    raise RuntimeError("统计表与图层数据不一致，已终止。确认继续时请重新运行并选择继续。")


def grouped_stats_by_soil(rows: Iterable[Dict], excludes: List[str]) -> OrderedDict:
    """按土类对数据进行分组，并自动对相同的【亚类+土属】组合进行面积合并累加与去重。"""
    excl = set(excludes)
    grouped: OrderedDict = OrderedDict()
    agg_track: OrderedDict = OrderedDict()

    for row in rows:
        sc = str(row["土类"]).strip()
        if sc and sc not in excl:
            if sc not in agg_track:
                agg_track[sc] = OrderedDict()

            subclass = str(row["亚类"]).strip()
            genus = str(row["土属"]).strip()
            k = (subclass, genus)

            if k not in agg_track[sc]:
                agg_track[sc][k] = {
                    "土类": sc,
                    "亚类": subclass,
                    "土属": genus,
                    "面积_亩": 0.0
                }
            agg_track[sc][k]["面积_亩"] += float(row.get("面积_亩", 0.0))

    for sc, k_dict in agg_track.items():
        grouped[sc] = list(k_dict.values())

    return grouped


def build_full_subclass_stats(
        stats_by_soil: Dict[str, List[Dict]],
        soil_classes: List[str],
) -> List[Dict]:
    """将全部土类的统计行聚合到亚类级别，对相同的【土类+亚类】组合进行面积去重累加。"""
    agg: OrderedDict = OrderedDict()
    for sc in soil_classes:
        for row in stats_by_soil.get(sc, []):
            k = (str(row["土类"]).strip(), str(row["亚类"]).strip())
            if k not in agg:
                agg[k] = {"土类": row["土类"], "亚类": row["亚类"], "面积_亩": 0.0}
            agg[k]["面积_亩"] += float(row["面积_亩"])
    return list(agg.values())


def span_ranges(rows: List[Dict], key: str) -> List[Tuple[int, int, str]]:
    """计算列表中某字段值连续相同的行范围（用于 Excel 合并单元格）。"""
    if not rows:
        return []
    spans: List[Tuple[int, int, str]] = []
    start = 0
    cur = str(rows[0].get(key, ""))
    for i, row in enumerate(rows[1:], 1):
        v = str(row.get(key, ""))
        if v != cur:
            spans.append((start, i - 1, cur))
            start, cur = i, v
    spans.append((start, len(rows) - 1, cur))
    return spans


def safe_filename(name: str) -> str:
    for ch in r'\/:*?"<>|':
        name = name.replace(ch, "_")
    return name.strip() or "未命名"


def make_layout_layer_name(layout_title: str) -> str:
    return safe_filename(f"{layout_title}_{LAYER_NAME_SUFFIX}")


def make_full_domain_layer_name(layout_title: str) -> str:
    return safe_filename(f"{layout_title}_全域数据_{LAYER_NAME_SUFFIX}")


def resolve_export_aprx(output_dir: Path, aprx_path: str = "") -> Path:
    """第二步导出时解析要使用的 .aprx。"""
    if aprx_path:
        p = Path(aprx_path).resolve()
        if not p.exists():
            raise RuntimeError(f"未找到布局工程：{p}")
        return p

    output_dir = Path(output_dir).resolve()
    default_aprx = output_dir / "自动化出图工作空间.aprx"
    if default_aprx.exists():
        return default_aprx.resolve()

    candidates = sorted(output_dir.glob("*.aprx"))
    if len(candidates) == 1:
        return candidates[0].resolve()
    if not candidates:
        raise RuntimeError(
            f"成果输出目录中未找到 .aprx 工程文件：{output_dir}。"
            "请确认该目录内已有保存好的布局工程。"
        )
    names = "、".join(p.name for p in candidates)
    raise RuntimeError(
        f"成果输出目录中发现多个 .aprx 工程文件：{names}。"
        "请保留一个要导出的工程，或在命令行使用 --aprx 明确指定。"
    )


def sql_eq(arcpy, fc: str, field: str, value: str) -> str:
    d = arcpy.AddFieldDelimiters(fc, field)
    return f"{d} = '{value.replace(chr(39), chr(39) * 2)}'"


def sql_not_in(arcpy, fc: str, field: str, values: List[str]) -> str:
    d = arcpy.AddFieldDelimiters(fc, field)
    quoted = ", ".join(f"'{v.replace(chr(39), chr(39)*2)}'" for v in values)
    return f"{d} NOT IN ({quoted})"


def detect_county(text: str) -> str:
    """从路径或文件名中提取县/区/市名（取最后一个匹配）。"""
    matches = re.findall(r"[\u4e00-\u9fa5]{2,10}(?:县|区|市|自治县)", str(text))
    return matches[-1] if matches else "XX县"


# =============================================================================
# 第五部分：ArcGIS Pro 布局操作工具
# =============================================================================

def validate_fields(arcpy, fc: str, fields: List[str]) -> None:
    existing = {f.name for f in arcpy.ListFields(fc)}
    missing = [f for f in fields if f and f not in existing]
    if missing:
        raise RuntimeError(
            f"字段不存在：{', '.join(missing)}；可用字段：{', '.join(sorted(existing))}"
        )


def get_blank_aprx(arcpy) -> Path:
    """获取 ArcGIS Pro 内置空白工程模板路径。"""
    candidates: List[Path] = []
    try:
        install_dir = Path(arcpy.GetInstallInfo()["InstallDir"])
        candidates.append(
            install_dir / "Resources/ArcToolBox/Services/routingservices/data/Blank.aprx"
        )
    except Exception:
        pass
    candidates.append(
        Path(r"C:\Program Files\ArcGIS\Pro\Resources\ArcToolBox"
             r"\Services\routingservices\data\Blank.aprx")
    )
    for p in candidates:
        if p.exists():
            return p
    tried = "\n  ".join(str(p) for p in candidates)
    raise RuntimeError(
        f"找不到 ArcGIS Pro 空白模板（.aprx），已尝试：\n  {tried}\n"
        "请确认 ArcGIS Pro 安装目录，或在 get_blank_aprx() 中手动指定路径。"
    )


def get_or_add_layer(arcpy, map_obj, source_layer: str, layer_name: str):
    source_name = Path(source_layer).name
    _quarantine_template_layers(map_obj, {layer_name, source_name})
    lyr = map_obj.addDataFromPath(str(Path(source_layer).resolve()))
    lyr.name = layer_name
    return lyr


def _quarantine_template_layers(map_obj, names) -> None:
    targets = {str(n) for n in names if str(n)}
    if not targets:
        return
    idx = 1
    for lyr in map_obj.listLayers():
        try:
            if lyr.name in targets:
                old_name = lyr.name
                lyr.name = f"{old_name}_模板旧图层_{idx}"
                idx += 1
        except Exception:
            pass


def add_layer_file(arcpy, map_obj, layer_file: Path, layer_name: str, source_layer: str = ""):
    lf = arcpy.mp.LayerFile(str(layer_file))
    _quarantine_template_layers(map_obj, {layer_name})
    map_obj.addLayer(lf)
    source_norm = _normalize_layer_source(source_layer)
    for lyr in map_obj.listLayers():
        if not _is_symbolizable_feature_layer(lyr):
            continue
        if lyr.name == layer_name:
            return lyr
        if source_norm and _normalize_layer_source(getattr(lyr, "dataSource", "")) == source_norm:
            lyr.name = layer_name
            return lyr
    available = [getattr(lyr, "name", "") for lyr in map_obj.listLayers()]
    raise RuntimeError(f"从样式图层文件添加图层失败：{layer_file}；当前图层：{available}")


def _normalize_layer_source(path_text: str) -> str:
    if not path_text:
        return ""
    text = str(path_text).replace("/", "\\").rstrip("\\")
    try:
        text = str(Path(text).resolve()).replace("/", "\\").rstrip("\\")
    except Exception:
        pass
    return text.casefold()


def _layer_supports(layer, capability: str) -> bool:
    try:
        supports = getattr(layer, "supports", None)
        if callable(supports):
            return bool(supports(capability))
    except Exception:
        return False
    return False


def _is_symbolizable_feature_layer(layer) -> bool:
    try:
        if not bool(getattr(layer, "isFeatureLayer", False)):
            return False
    except Exception:
        return False
    return _layer_supports(layer, "SYMBOLOGY")


def find_main_map_frame(layout):
    """从布局中找到主底图地图框（面积最大的 MAPFRAME_ELEMENT）。"""
    map_frames = layout.listElements("MAPFRAME_ELEMENT")
    if not map_frames:
        raise RuntimeError(
            f"布局「{layout.name}」中未找到任何地图框（MAPFRAME_ELEMENT），"
            "请确认模板文件包含至少一个地图框。"
        )
    main_mf = max(map_frames, key=lambda f: f.elementWidth * f.elementHeight)
    if len(map_frames) > 1:
        others = [f.name for f in map_frames if f is not main_mf]
        print(
            f"  布局「{layout.name}」共检测到 {len(map_frames)} 个地图框，"
            f"已选主底图框「{main_mf.name}」（最大），"
            f"忽略位置示意图框：{others}。"
        )
    return main_mf


def find_county_title(layout, county_name: str = ""):
    """定位模板中"图名"文本元素。"""
    _noise = {"比例尺", "图例", "指北针", "普查", "审图", "制图", "坐标", "投影"}

    if county_name:
        for e in layout.listElements("TEXT_ELEMENT"):
            if (e.text or "").strip() == county_name:
                return e

    for e in layout.listElements("TEXT_ELEMENT"):
        t = (e.text or "").strip()
        if re.fullmatch(r"[\u4e00-\u9fa5]{2,10}(?:县|区|市|自治县)", t):
            return e

    for e in layout.listElements("TEXT_ELEMENT"):
        t = (e.text or "").strip()
        if t and re.search(r"[\u4e00-\u9fa5]{2,10}(?:县|区|市|自治县)", t):
            if not any(kw in t for kw in _noise):
                return e

    return None


def _layout_key(layout) -> Tuple[str, str]:
    return (str(getattr(layout, "URI", "") or ""), str(getattr(layout, "name", "") or ""))


def import_layout(aprx, mxd_path: str, layout_title: str):
    """导入 MXD 后精确返回新增布局。"""
    before_keys = {_layout_key(layout) for layout in aprx.listLayouts()}
    before_names = defaultdict(int)
    for _, name in before_keys:
        before_names[name] += 1

    aprx.importDocument(mxd_path)

    after = aprx.listLayouts()
    new_layouts = [layout for layout in after if _layout_key(layout) not in before_keys]
    if not new_layouts:
        after_names = defaultdict(int)
        for layout in after:
            after_names[str(getattr(layout, "name", "") or "")] += 1
        new_names = [name for name, count in after_names.items() if count > before_names.get(name, 0)]
        new_layouts = [layout for layout in after if str(getattr(layout, "name", "") or "") in new_names]

    if not new_layouts:
        raise RuntimeError(f"导入模板后未检测到新增布局：{mxd_path}")
    if len(new_layouts) > 1:
        names = "、".join(str(getattr(layout, "name", "") or "") for layout in new_layouts)
        print(f"警告：导入模板后检测到多个新增布局：{names}；将使用第一个布局。", file=sys.stderr)

    layout = new_layouts[0]
    layout.name = safe_filename(layout_title)
    return layout


def suppress_legend_sync(layout) -> list:
    states = []
    for leg in layout.listElements("LEGEND_ELEMENT"):
        try:
            states.append((leg, leg.syncNewLayer))
            leg.syncNewLayer = False
        except Exception:
            pass
    return states


def restore_legend_sync(states: list) -> None:
    for leg, s in states:
        try:
            leg.syncNewLayer = s
        except Exception:
            pass


def _refresh_layer(mf, layer_name: str, source_layer: str = ""):
    """
    从指定地图框的 map 中重新搜索并返回有效图层引用。

    ArcGIS Pro 在 importDocument / addDataFromPath 之后，原始返回的 Layer
    对象可能因内部指针失效而丢失 symbology 等属性（表现为
    "The attribute 'symbology' is not supported on this instance of Layer"）。
    必须通过 mf.map.listLayers() 重新捞取，才能得到与当前布局地图框真正
    绑定的、完整的 FeatureLayer 实例。
    """
    source_norm = _normalize_layer_source(source_layer)
    candidates = []
    for lyr in mf.map.listLayers():
        if _is_symbolizable_feature_layer(lyr) and lyr.name == layer_name:
            candidates.append(lyr)
            if not source_norm:
                return lyr
            if _normalize_layer_source(getattr(lyr, "dataSource", "")) == source_norm:
                return lyr
    available = [l.name for l in mf.map.listLayers()]
    candidate_sources = [
        getattr(lyr, "dataSource", "") for lyr in candidates
    ]
    raise RuntimeError(
        f"在地图框「{mf.name}」对应的地图中未找到可符号化的要素图层「{layer_name}」。\n"
        f"目标数据源：{source_layer or '未指定'}\n"
        f"同名候选数据源：{candidate_sources}\n"
        f"当前图层列表：{available}"
    )


def _set_layer_symbology(layer, sym, stage: str, refresh_layer=None):
    try:
        layer.symbology = sym
        return layer
    except Exception as first_error:
        if refresh_layer is not None:
            try:
                fresh_layer = refresh_layer()
                fresh_layer.symbology = sym
                return fresh_layer
            except Exception as second_error:
                raise RuntimeError(
                    f"{stage}失败；刷新图层后仍无法写入 symbology。"
                    f"原图层 name={getattr(layer, 'name', '')!r}, "
                    f"class={layer.__class__.__name__}, "
                    f"supports_SYMBOLOGY={_layer_supports(layer, 'SYMBOLOGY')!r}；"
                    f"第一次错误：{first_error}；第二次错误：{second_error}"
                ) from second_error
        raise RuntimeError(
            f"{stage}失败；当前 Layer 实例不允许写入 symbology。"
            f"name={getattr(layer, 'name', '')!r}, "
            f"class={layer.__class__.__name__}, "
            f"supports_SYMBOLOGY={_layer_supports(layer, 'SYMBOLOGY')!r}；"
            f"ArcGIS 原始错误：{first_error}"
        ) from first_error


def _make_cim_solid_fill_symbol(rgb: Tuple[int, int, int]) -> dict:
    """构造一个纯色填充、无描边的 CIM 面符号（JSON dict）。"""
    return {
        "type": "CIMPolygonSymbol",
        "symbolLayers": [
            {
                "type": "CIMSolidFill",
                "enable": True,
                "color": {
                    "type": "CIMRGBColor",
                    "values": [rgb[0], rgb[1], rgb[2], 100],
                },
            },
            {
                "type": "CIMSolidStroke",
                "enable": False,
                "width": 0,
                "color": {"type": "CIMRGBColor", "values": [0, 0, 0, 0]},
            },
        ],
    }


def _make_cim_transparent_symbol() -> dict:
    """构造全透明（不显示）的 CIM 面符号（JSON dict），用于无颜色映射的值。"""
    return {
        "type": "CIMPolygonSymbol",
        "symbolLayers": [
            {
                "type": "CIMSolidFill",
                "enable": False,
                "color": {"type": "CIMRGBColor", "values": [0, 0, 0, 0]},
            },
            {
                "type": "CIMSolidStroke",
                "enable": False,
                "width": 0,
                "color": {"type": "CIMRGBColor", "values": [0, 0, 0, 0]},
            },
        ],
    }


def _build_cim_unique_value_renderer(
        render_field: str,
        value_colors: Dict[str, Tuple[int, int, int]],
        extra_values: List[str],
) -> dict:
    """
    纯 JSON 构造 CIMUniqueValueRenderer，完全不依赖 layer.symbology 赋值。

    extra_values：图层中实际存在但 value_colors 里没有颜色的字段值（透明处理）。
    """
    classes = []
    for val, rgb in value_colors.items():
        classes.append({
            "type": "CIMUniqueValueClass",
            "label": val,
            "patch": "Default",
            "symbol": {
                "type": "CIMSymbolReference",
                "symbol": _make_cim_solid_fill_symbol(rgb),
            },
            "values": [
                {
                    "type": "CIMUniqueValue",
                    "fieldValues": [val],
                }
            ],
            "visible": True,
        })
    for val in extra_values:
        if val not in value_colors:
            classes.append({
                "type": "CIMUniqueValueClass",
                "label": val,
                "patch": "Default",
                "symbol": {
                    "type": "CIMSymbolReference",
                    "symbol": _make_cim_transparent_symbol(),
                },
                "values": [
                    {
                        "type": "CIMUniqueValue",
                        "fieldValues": [val],
                    }
                ],
                "visible": True,
            })

    return {
        "type": "CIMUniqueValueRenderer",
        "defaultLabel": "<all other values>",
        "defaultSymbol": {
            "type": "CIMSymbolReference",
            "symbol": _make_cim_transparent_symbol(),
        },
        "defaultSymbolPatch": "Default",
        "fields": [render_field],
        "groups": [
            {
                "type": "CIMUniqueValueGroup",
                "classes": classes,
                "heading": render_field,
            }
        ],
        "useDefaultSymbol": False,
        "polygonSymbolColorTarget": "Fill",
    }


def apply_cim_symbology(
        layer,
        render_field: str,
        value_colors: Dict[str, Tuple[int, int, int]],
        refresh_layer=None,
) -> None:
    """
    通过纯 CIM getDefinition/setDefinition 应用唯一值配色。

    完全不使用 layer.symbology = ... 赋值，彻底规避 ArcGIS Pro 在
    多布局 importDocument 场景下 Layer 对象指针失效导致
    "The attribute 'symbology' is not supported on this instance of Layer" 的问题。

    流程：
    1. 尝试从现有 CIM 里读取已有的唯一值字段值列表（作为 extra_values）
    2. 用纯 JSON 构造完整的 CIMUniqueValueRenderer
    3. 替换 CIM 中的 renderer 字段后写回 setDefinition
    """
    if not _is_symbolizable_feature_layer(layer):
        # 尝试刷新一次
        if refresh_layer is not None:
            try:
                layer = refresh_layer()
            except Exception:
                pass
        if not _is_symbolizable_feature_layer(layer):
            raise RuntimeError(
                f"图层「{getattr(layer, 'name', '')}」未能识别为可符号化要素图层，"
                f"class={layer.__class__.__name__}, "
                f"supports_SYMBOLOGY={_layer_supports(layer, 'SYMBOLOGY')!r}"
            )

    # ── 读取现有 CIM，收集图层中实际有的字段值（用于透明处理未配色的值）────
    extra_values: List[str] = []
    try:
        cim = layer.getDefinition("V2")
        existing_renderer = getattr(cim, "renderer", None)
        if existing_renderer is not None:
            for grp in getattr(existing_renderer, "groups", []) or []:
                for cls in getattr(grp, "classes", []) or []:
                    v = _extract_class_value(cls)
                    if v and v not in value_colors and v not in extra_values:
                        extra_values.append(v)
    except Exception as e:
        print(f"提示：读取现有渲染器失败（{e}），将直接构造新渲染器。", file=sys.stderr)
        try:
            cim = layer.getDefinition("V2")
        except Exception as e2:
            raise RuntimeError(
                f"getDefinition 失败，无法应用唯一值配色：{e2}\n"
                f"name={getattr(layer, 'name', '')!r}, "
                f"class={layer.__class__.__name__}"
            ) from e2

    # ── 构造并写入新渲染器 ────────────────────────────────────────────────────
    new_renderer = _build_cim_unique_value_renderer(render_field, value_colors, extra_values)
    cim.renderer = new_renderer
    layer.setDefinition(cim)


def apply_simple_fill_cim(layer, rgb: Tuple[int, int, int]) -> None:
    cim = layer.getDefinition("V2")
    renderer = getattr(cim, "renderer", None)
    symbol_ref = getattr(renderer, "symbol", None)
    symbol = getattr(symbol_ref, "symbol", None)
    if symbol is None:
        raise RuntimeError(f"图层「{getattr(layer, 'name', '')}」没有可修改的简单符号。")
    for sl in getattr(symbol, "symbolLayers", []) or []:
        n = sl.__class__.__name__
        if n == "CIMSolidFill":
            sl.enable = True
            sl.color = {"type": "CIMRGBColor", "values": [rgb[0], rgb[1], rgb[2], 100]}
        elif n == "CIMSolidStroke":
            sl.enable = False
            sl.width = 0
            sl.color = {"type": "CIMRGBColor", "values": [0, 0, 0, 0]}
    layer.setDefinition(cim)


def sql_and(*parts: str) -> str:
    clean = [f"({p})" for p in parts if p]
    return " AND ".join(clean)


def _extract_class_value(cls) -> str:
    if hasattr(cls, "values") and cls.values:
        v = cls.values[0]
        if hasattr(v, "fieldValues") and v.fieldValues:
            return str(v.fieldValues[0]).strip()
        if hasattr(v, "value"):
            return str(v.value).strip()
    return str(getattr(cls, "label", "")).strip()


def _clear_outline(layer) -> None:
    try:
        cim = layer.getDefinition("V2")
        for sr in filter(None, [
            getattr(cim.renderer, "defaultSymbol", None),
            getattr(cim.renderer, "symbol", None),
        ]):
            sym = getattr(sr, "symbol", sr)
            for sl in getattr(sym, "symbolLayers", []) or []:
                if sl.__class__.__name__ == "CIMSolidStroke":
                    sl.enable = False
                    sl.width = 0
        layer.setDefinition(cim)
    except Exception:
        pass


def safe_layer_file(arcpy, layer, out_path: Path) -> None:
    try:
        arcpy.management.SaveToLayerFile(layer, str(out_path), "RELATIVE")
        print(f"样式文件已保存 -> {out_path.name}")
    except Exception as e:
        print(f"警告：.lyrx 保存失败：{e}", file=sys.stderr)


# =============================================================================
# 【核心修改】合并配色：单图层唯一值渲染
# =============================================================================

def add_single_layer_with_unique_values(
        arcpy,
        map_obj,
        mf,
        source_layer: str,
        layer_name: str,
        definition_query: str,
        render_field: str,
        value_colors: Dict[str, Tuple[int, int, int]],
) -> object:
    """
    添加单个图层，设置定义查询后，用唯一值渲染器（CIM）一次性着色所有分类值。

    与旧版 add_simple_value_layers 的区别：
    - 旧版：每个分类值（亚类/土属）创建一个独立图层 → N 个土属 = N 个图层
    - 新版：整个土类只创建 1 个图层，所有分类值用唯一值渲染器在同一图层内区分颜色

    参数
    ----
    arcpy           : arcpy 模块
    map_obj         : arcpy.mp.Map 对象
    mf              : 布局中的 MapFrame 对象（用于 _refresh_layer 重新捞取指针）
    source_layer    : 要素图层路径
    layer_name      : 添加后的图层显示名称
    definition_query: 图层定义查询（如 "TL = '红壤'"），空字符串表示不过滤
    render_field    : 唯一值渲染字段（土属字段或亚类字段）
    value_colors    : {字段值: (R, G, B)} 颜色映射字典
    """
    if not value_colors:
        raise RuntimeError(f"图层「{layer_name}」没有可渲染的分类值，请检查统计表与配色表。")

    # ── 1. 添加图层，隔离同名旧图层 ────────────────────────────────────────
    lyr = get_or_add_layer(arcpy, map_obj, source_layer, layer_name)

    # ── 2. 设置定义查询（仅显示本土类要素）──────────────────────────────────
    if definition_query:
        lyr.definitionQuery = definition_query

    # ── 3. 重新从地图框捞取完整 Layer 实例（防止 Pro 内存指针失效）──────────
    #       importDocument 之后原始 lyr 对象的 symbology 属性常常不可写，
    #       必须通过 mf.map.listLayers() 重新获取。
    def do_refresh():
        return _refresh_layer(mf, layer_name, source_layer)

    try:
        lyr = do_refresh()
    except Exception as e:
        print(f"警告：首次刷新图层引用失败（{e}），将使用原始引用继续尝试。", file=sys.stderr)

    # ── 4. 应用唯一值 CIM 配色 ───────────────────────────────────────────────
    apply_cim_symbology(lyr, render_field, value_colors, refresh_layer=do_refresh)

    return lyr


# =============================================================================
# 第六部分：多表单 Excel 配色对照表导出
# =============================================================================

def export_multi_sheet_excel(
        path: Path,
        soil_classes: List[str],
        stats_by_soil: Dict[str, List[Dict]],
        full_sub_stats: List[Dict],
        full_sub_colors: Dict[str, Tuple[int, int, int]],
        palettes: Dict[str, List[Tuple[int, int, int]]],
        area_unit: str = DEFAULT_AREA_UNIT,
        color_direction: str = DEFAULT_COLOR_DIRECTION,
) -> None:
    """导出多表单配色对照 Excel。"""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        print("未检测到 openpyxl，跳过 Excel 导出。", file=sys.stderr)
        return

    area_label = AREA_UNIT_LABELS.get(area_unit, AREA_UNIT_LABELS[DEFAULT_AREA_UNIT])
    area_factor = AREA_UNIT_FACTORS.get(area_unit, 1.0)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    hdr_f = Font(name="宋体", size=11, bold=True)
    txt_f = Font(name="宋体", size=11)
    num_f = Font(name="Times New Roman", size=11)
    white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    def _write_sheet(
            ws,
            headers: List[str],
            color_col: int,
            color_key: str,
            rows: List[Dict],
            vc: Dict[str, Tuple[int, int, int]],
            merge_soil_by_span: bool = False,
    ) -> None:
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_f
            c.border = border
            c.alignment = center
            c.fill = white
        ws.row_dimensions[1].height = 26
        data_start = 2

        for ri, item in enumerate(rows, data_start):
            ws.row_dimensions[ri].height = 22
            row_vals = [
                float(item.get("面积_亩", 0)) * area_factor if h == area_label
                else item.get(h, "")
                for h in headers
            ]
            target = str(item.get(color_key, "")).strip()
            rgb = vc.get(target, (255, 255, 255))
            cfill = PatternFill(
                start_color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}",
                end_color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}",
                fill_type="solid",
            )
            for ci, v in enumerate(row_vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = border
                cell.alignment = center
                cell.font = num_f if ci == len(row_vals) else txt_f
                if ci == len(row_vals):
                    # 亩：整数；万亩：两位小数；均不使用千分位分隔符
                    cell.number_format = "0.00" if area_unit == "万亩" else "0"
                cell.fill = cfill if ci == color_col else white

        data_end = data_start + len(rows) - 1
        if data_end < data_start:
            return

        if merge_soil_by_span:
            for s, e, _ in span_ranges(rows, "土类"):
                if e > s:
                    ws.merge_cells(
                        start_row=data_start + s, start_column=1,
                        end_row=data_start + e, end_column=1,
                    )
        else:
            ws.merge_cells(
                start_row=data_start, start_column=1,
                end_row=data_end, end_column=1,
            )

        if len(headers) == 4:
            for s, e, _ in span_ranges(rows, "亚类"):
                if e > s:
                    ws.merge_cells(
                        start_row=data_start + s, start_column=2,
                        end_row=data_start + e, end_column=2,
                    )

        for col in ws.columns:
            ml = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col[0].column)
            ].width = max(ml + 4, 13)

    for sc in soil_classes:
        block = stats_by_soil.get(sc, [])
        if not block:
            continue
        vc = build_value_colors(sc, block, palettes, use_subclass_mode=False,
                                color_direction=color_direction)
        ws = wb.create_sheet(title=safe_filename(sc)[:31])
        _write_sheet(ws, ["土类", "亚类", "土属", area_label], 3, "土属", block, vc,
                     merge_soil_by_span=False)

    ws_full = wb.create_sheet(title="全域亚类配色")
    _write_sheet(ws_full, ["土类", "亚类", area_label], 2, "亚类",
                 full_sub_stats, full_sub_colors, merge_soil_by_span=True)

    wb.save(str(path))
    print(f"配色对照表已保存 -> {path}")


# =============================================================================
# 第七部分：主出图工作流
# =============================================================================

def _setup_layout(
        aprx,
        arcpy,
        mxd_path: str,
        layout_title: str,
        map_title: str,
        county_name: str = "",
) -> Tuple:
    """
    从 mxd 导入一个新布局并完成基础配置（图名更新）。

    注意：不再在此处添加图层，图层由调用方在获得布局和地图框后单独添加，
    以确保能向 add_single_layer_with_unique_values 传递有效的 mf 引用。

    返回 (layout, mf, template_scale)
    """
    layout = import_layout(aprx, mxd_path, layout_title)
    mf = find_main_map_frame(layout)
    template_scale = mf.camera.scale

    try:
        mo = mf.map
    except Exception:
        mo = aprx.listMaps()[-1]
    mf.map = mo

    title_elm = find_county_title(layout, county_name)
    if title_elm:
        title_elm.text = map_title
    else:
        print(
            f"警告：布局「{layout_title}」中未找到县/市名称文本元素，图名未能自动更新。",
            file=sys.stderr,
        )

    return layout, mf, template_scale


def export_with_arcgis_pro(args) -> Dict[str, Path]:
    """主出图函数。"""
    import arcpy  # type: ignore  # 延迟导入，GUI 启动时不加载
    _ensure_openpyxl()
    if not hasattr(arcpy, "mp"):
        raise RuntimeError("当前环境无 arcpy.mp，请使用 ArcGIS Pro 的 Python 环境。")

    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    # ── 第二步：批量导出 JPG ──────────────────────────────────────────────────
    if args.mode == "export":
        work_aprx = resolve_export_aprx(output_dir, getattr(args, "aprx", ""))
        aprx = arcpy.mp.ArcGISProject(str(work_aprx))
        layouts = aprx.listLayouts()
        print(f"从工程 {work_aprx.name} 检测到 {len(layouts)} 个布局，开始批量导出...")
        out: Dict[str, Path] = {}
        for layout in layouts:
            if layout.name in ("Layout", "布局"):
                continue
            jpg = (output_dir / layout.name).with_suffix(".jpg")
            layout.exportToJPEG(str(jpg), resolution=args.dpi)
            out[layout.name] = jpg
            print(f"已导出 -> {jpg.name}")
        return out

    # ── 第一步：生成多布局工程 ────────────────────────────────────────────────
    work_aprx = output_dir / "自动化出图工作空间.aprx"
    mxd_path = str(Path(args.mxd).resolve())
    source_layer = str(Path(args.source_layer).resolve())
    stats_path = Path(args.stats_table).resolve()
    print(f"脚本文件 -> {Path(__file__).resolve()}", flush=True)
    print("脚本版本 -> 2026-05-31-single-layer-unified-renderer", flush=True)

    if not Path(mxd_path).exists():
        raise RuntimeError(f"找不到图框模板文件：{mxd_path}")
    if not arcpy.Exists(source_layer):
        raise RuntimeError(f"找不到要素图层：{source_layer}")

    county = (getattr(args, "county", "") or "").strip()
    if not county:
        county = detect_county(source_layer) or detect_county(mxd_path)
    genus_field = (args.genus_field or "").strip()
    render_genus = genus_field if genus_field else args.subclass_field
    check_fields = [args.soil_field, args.subclass_field] + ([genus_field] if genus_field else [])
    validate_fields(arcpy, source_layer, check_fields)

    excludes = split_names(args.excludes)
    palettes = read_color_table(Path(args.color_table).resolve())
    raw_rows = read_stats_table(stats_path)
    stats_by_soil = grouped_stats_by_soil(raw_rows, excludes)
    layer_type_sets = list_layer_type_sets(
        arcpy, source_layer, args.soil_field, args.subclass_field, genus_field, excludes,
    )
    mismatch_report = build_type_mismatch_report(
        stats_by_soil, layer_type_sets, compare_genus=bool(genus_field),
    )
    confirm_type_mismatch(mismatch_report, args)

    layer_soils = set(layer_type_sets["soil"])
    soil_classes = [s for s in stats_by_soil if s in layer_soils]

    if not soil_classes:
        raise RuntimeError("无可出图土类，请检查配色表/统计表与图层字段是否一致。")

    full_sub_stats = build_full_subclass_stats(stats_by_soil, soil_classes)
    full_sub_colors = build_full_subclass_colors(soil_classes, stats_by_soil, palettes,
                                                    color_direction=args.color_direction)

    blank = get_blank_aprx(arcpy)
    print(f"阶段 -> 找到空白工程 {blank}", flush=True)
    shutil.copy2(str(blank), str(work_aprx))
    print(f"阶段 -> 已复制工作工程 {work_aprx}", flush=True)
    aprx = arcpy.mp.ArcGISProject(str(work_aprx))
    print("阶段 -> 已打开工作工程", flush=True)

    full_title = f"{county}{FULL_DOMAIN_SUFFIX}"
    full_layer_name = make_full_domain_layer_name(full_title)

    # ── 按土类生成布局（每布局仅 1 个图层，唯一值渲染土属/亚类颜色）──────────
    for sc in soil_classes:
        # use_subclass_mode=True 当无土属字段时按亚类配色，否则按土属配色
        vc = build_value_colors(sc, stats_by_soil[sc], palettes, use_subclass_mode=not genus_field,
                                color_direction=args.color_direction)
        sub_title = f"{county}{sc}分布图"
        layer_name = make_layout_layer_name(sub_title)

        # 导入布局、获取地图框，不添加图层
        layout, mf, template_scale = _setup_layout(
            aprx, arcpy, mxd_path, sub_title, sub_title, county,
        )

        # 构建定义查询：仅显示本土类要素
        soil_query = sql_eq(arcpy, source_layer, args.soil_field, sc)

        states = suppress_legend_sync(layout)
        try:
            # ★ 核心改动：整个土类 = 1 个图层 + 唯一值渲染器
            add_single_layer_with_unique_values(
                arcpy=arcpy,
                map_obj=mf.map,
                mf=mf,
                source_layer=source_layer,
                layer_name=layer_name,
                definition_query=soil_query,
                render_field=render_genus,
                value_colors=vc,
            )
        finally:
            restore_legend_sync(states)

        mf.camera.scale = template_scale
        print(f"布局装载 -> {layout.name}（1 个图层，{len(vc)} 个唯一值）")

    # ── 全域亚类配色布局（1 个图层，唯一值渲染全部亚类颜色）────────────────
    print(f"全域内部图层名 -> {full_layer_name}")
    layout_full, mf_full, template_scale_full = _setup_layout(
        aprx, arcpy, mxd_path, full_title, full_title, county,
    )

    # 构建全域定义查询：排除建设用地等
    if excludes:
        full_base_query = sql_not_in(arcpy, source_layer, args.soil_field, excludes)
    else:
        full_base_query = ""

    states_full = suppress_legend_sync(layout_full)
    try:
        # ★ 核心改动：全域亚类也是 1 个图层 + 唯一值渲染器
        add_single_layer_with_unique_values(
            arcpy=arcpy,
            map_obj=mf_full.map,
            mf=mf_full,
            source_layer=source_layer,
            layer_name=full_layer_name,
            definition_query=full_base_query,
            render_field=args.subclass_field,
            value_colors=full_sub_colors,
        )
    finally:
        restore_legend_sync(states_full)

    mf_full.camera.scale = template_scale_full
    print(f"全域布局装载 -> {layout_full.name}（1 个图层，{len(full_sub_colors)} 个唯一值）")

    excel_path = output_dir / f"{county}{LAYER_NAME_SUFFIX}.xlsx"
    export_multi_sheet_excel(
        excel_path, soil_classes, stats_by_soil,
        full_sub_stats, full_sub_colors, palettes,
        area_unit=args.area_unit,
        color_direction=args.color_direction,
    )

    aprx.save()
    return {"aprx": work_aprx, "excel": excel_path}


# =============================================================================
# 第八部分：Tkinter 图形界面
# =============================================================================

# GUI 界面配置持久化：记忆用户上次填写的各路径/字段，下次启动自动回填。
# 存到用户目录而非脚本目录——单文件 EXE 运行时脚本位于临时解压目录，
# 用 LOCALAPPDATA 才能跨次稳定保存。
GUI_CONFIG_KEYS = (
    "mxd", "source_layer", "color_table", "stats_table", "output_dir",
    "county", "excludes", "fields", "area_unit", "color_direction",
)


def gui_config_path() -> Path:
    base = os.environ.get("LOCALAPPDATA") or os.environ.get("APPDATA") or str(Path.home())
    return Path(base) / "土壤类型出图工具" / "界面配置.json"


def load_gui_config() -> Dict[str, str]:
    """读取上次保存的界面配置，失败时返回空字典（不影响启动）。"""
    try:
        p = gui_config_path()
        if p.exists():
            data = json.loads(p.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return {k: str(v) for k, v in data.items() if k in GUI_CONFIG_KEYS}
    except Exception as e:
        print(f"提示：读取界面配置失败（{e}），使用默认值。", file=sys.stderr)
    return {}


def save_gui_config(data: Dict[str, str]) -> None:
    """保存界面配置，失败仅警告不报错。"""
    try:
        p = gui_config_path()
        p.parent.mkdir(parents=True, exist_ok=True)
        payload = {k: str(data.get(k, "")) for k in GUI_CONFIG_KEYS}
        p.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        print(f"提示：保存界面配置失败（{e}）。", file=sys.stderr)


def run_gui() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox, simpledialog

    root = tk.Tk()
    root.title("土壤类型图智能批量出图工具")
    root.geometry("900x540")

    vars_ = {
        "mxd": tk.StringVar(value=str(CFG_FRAMES_BASE_DIR)),
        "source_layer": tk.StringVar(value=str(CFG_SOIL_GDB_BASE_DIR)),
        "color_table": tk.StringVar(value=str(CFG_COLOR_TABLE)),
        "stats_table": tk.StringVar(value=str(CFG_STATS_BASE_DIR)),
        "output_dir": tk.StringVar(value=str(CFG_OUTPUT_BASE_DIR)),
        "county": tk.StringVar(value=""),
        "excludes": tk.StringVar(value=DEFAULT_EXCLUDES),
        "fields": tk.StringVar(
            value=f"{DEFAULT_SOIL_FIELD},{DEFAULT_SUBCLASS_FIELD},{DEFAULT_GENUS_FIELD}"
        ),
        "area_unit": tk.StringVar(value=DEFAULT_AREA_UNIT),
        "color_direction": tk.StringVar(value="深→浅"),
    }

    # 回填上次保存的配置（仅覆盖有值的项，保留默认值兜底）。
    _saved = load_gui_config()
    for _k, _v in _saved.items():
        if _k in vars_ and _v:
            vars_[_k].set(_v)

    def snapshot_settings() -> Dict[str, str]:
        return {k: v.get() for k, v in vars_.items()}

    def make_row(label: str, key: str, browse=None, lbl_color: str = "black"):
        f = tk.Frame(root)
        f.pack(fill="x", padx=16, pady=4)
        tk.Label(f, text=label, width=14, anchor="w", fg=lbl_color).pack(side="left")
        tk.Entry(f, textvariable=vars_[key]).pack(side="left", fill="x", expand=True)
        if browse:
            tk.Button(f, text="选择", command=browse).pack(side="left", padx=6)

    def pick_file(key: str, ft: list):
        p = filedialog.askopenfilename(filetypes=ft)
        if p:
            vars_[key].set(str(Path(p).resolve()))

    def pick_dir(key: str):
        p = filedialog.askdirectory()
        if p:
            vars_[key].set(str(Path(p).resolve()))

    def pick_layer():
        gdb = filedialog.askdirectory(title="选择 .gdb 目录")
        if not gdb:
            return
        name = simpledialog.askstring("图层名称", "请输入 .gdb 内的目标图层名称")
        if name:
            vars_["source_layer"].set(str(Path(gdb).resolve() / name.strip()))

    make_row("A3 图框底图", "mxd",
             lambda: pick_file("mxd", [("MXD 模板", "*.mxd")]))
    make_row("土壤要素图层", "source_layer", pick_layer)

    _cf = tk.Frame(root)
    _cf.pack(fill="x", padx=16, pady=4)
    tk.Label(_cf, text="县/市名称", width=14, anchor="w").pack(side="left")
    tk.Entry(_cf, textvariable=vars_["county"], width=12).pack(side="left")
    tk.Label(
        _cf,
        text="  例：xx县  （留空则从数据路径自动识别；图名格式：xx县红壤分布图）",
        fg="#888888", font=("", 8),
    ).pack(side="left")

    make_row("配色表（固定）", "color_table",
             lambda: pick_file("color_table", [("CSV", "*.csv")]))
    make_row("平差面积统计", "stats_table",
             lambda: pick_file("stats_table", [("Excel", "*.xlsx")]))
    make_row("成果输出目录", "output_dir", lambda: pick_dir("output_dir"))
    make_row("剔除土类", "excludes")

    _opt = tk.Frame(root)
    _opt.pack(fill="x", padx=16, pady=4)
    tk.Label(_opt, text="面积单位", width=14, anchor="w").pack(side="left")
    tk.OptionMenu(_opt, vars_["area_unit"], "亩", "万亩").pack(side="left")
    tk.Label(_opt, text="    色带方向", anchor="w").pack(side="left", padx=(16, 0))
    _color_dir_map = {"深→浅": "dark_to_light", "浅→深": "light_to_dark"}
    tk.OptionMenu(_opt, vars_["color_direction"], *_color_dir_map.keys()).pack(side="left")
    tk.Label(
        _opt,
        text="  深→浅：从深色开始渐变为浅色；浅→深：从浅色开始渐变为深色",
        fg="#888888", font=("", 8),
    ).pack(side="left")

    ff = tk.Frame(root)
    ff.pack(fill="x", padx=16, pady=4)
    tk.Label(ff, text="字段映射", width=14, anchor="w").pack(side="left")
    tk.Entry(ff, textvariable=vars_["fields"]).pack(side="left", fill="x", expand=True)
    tk.Label(ff,
             text="  格式：土类,亚类,土属（留空土属则按亚类出图）",
             fg="#E67E22", font=("", 8, "bold")).pack(side="left")

    status = tk.StringVar(
        value="就绪。第一步可生成布局工程；已有布局工程时，设置成果输出目录后可直接按布局批量出图。"
    )
    tk.Label(root, textvariable=status, anchor="w", fg="#7F8C8D").pack(fill="x", padx=16, pady=6)

    def gui_confirm_dialog(text: str) -> bool:
        """线程安全的确认弹窗：在主线程显示 messagebox，子线程阻塞等待结果。"""
        answer: Dict[str, bool] = {}
        done = threading.Event()

        def _ask():
            try:
                answer["ok"] = messagebox.askyesno("土壤类型不一致", text)
            finally:
                done.set()

        root.after(0, _ask)
        done.wait()
        return answer.get("ok", False)

    def run(mode: str):
        save_gui_config(snapshot_settings())  # 每次执行前记忆当前填写
        parts = [p.strip() for p in vars_["fields"].get().replace("，", ",").split(",")]
        sf = parts[0] if len(parts) > 0 and parts[0] else DEFAULT_SOIL_FIELD
        yf = parts[1] if len(parts) > 1 and parts[1] else DEFAULT_SUBCLASS_FIELD
        gf = parts[2] if len(parts) > 2 else ""
        args = argparse.Namespace(
            mode=mode,
            county=vars_["county"].get(),
            mxd=vars_["mxd"].get(),
            source_layer=vars_["source_layer"].get(),
            color_table=vars_["color_table"].get(),
            stats_table=vars_["stats_table"].get(),
            output_dir=vars_["output_dir"].get(),
            aprx="",
            soil_field=sf,
            subclass_field=yf,
            genus_field=gf,
            excludes=vars_["excludes"].get(),
            area_unit=vars_["area_unit"].get(),
            color_direction=_color_dir_map.get(vars_["color_direction"].get(),
                                                DEFAULT_COLOR_DIRECTION),
            dpi=DEFAULT_DPI,
            confirm_mismatch=True,
            continue_on_mismatch=False,
            confirm_dialog=gui_confirm_dialog,
        )
        status.set("正在执行，请稍候...")

        def _worker():
            try:
                export_with_arcgis_pro(args)
                root.after(0, lambda: _on_done(mode, True, ""))
            except Exception as e:
                # 立即固化异常信息：except 块结束后 e 会被销毁，
                # 而 lambda 延迟到主线程执行时再引用 e 会报
                # "cannot access free variable 'e'"，反而掩盖真实错误。
                err_msg = str(e) or e.__class__.__name__
                traceback.print_exc(file=sys.stderr)
                root.after(0, lambda: _on_done(mode, False, err_msg))

        def _on_done(mode: str, ok: bool, err: str):
            if ok:
                if mode == "generate":
                    status.set("第一步完成：集中式工程已生成，请在 Pro 中微调后执行第二步。")
                    messagebox.showinfo(
                        "第一步成功",
                        "已生成【自动化出图工作空间.aprx】及配色对照 Excel。\n"
                        "请在 ArcGIS Pro 中打开工程，在布局列表中逐一微调版面，\n"
                        "保存后回到本工具点击【按照布局文件批量出图】。",
                    )
                else:
                    status.set(f"第二步完成：已批量导出 JPG 至 {vars_['output_dir'].get()}")
                    messagebox.showinfo(
                        "第二步成功",
                        f"所有布局已成功导出为高清 JPG：\n{vars_['output_dir'].get()}",
                    )
            else:
                status.set(f"执行出错：{err[:80]}")
                messagebox.showerror("执行出错", f"{err}\n\n详情请查看输出目录下的【出图日志.txt】。")

        t = threading.Thread(target=_worker, daemon=True)
        t.start()

    bf = tk.Frame(root)
    bf.pack(pady=10)
    tk.Button(
        bf, text="第一步：生成多布局工作空间 (.aprx)",
        command=lambda: run("generate"),
        bg="#3498DB", fg="white", font=("", 10, "bold"), height=2, padx=10,
    ).pack(side="left", padx=15)
    tk.Button(
        bf, text="第二步：按照布局文件批量出图（.jpg）",
        command=lambda: run("export"),
        bg="#2ECC71", fg="white", font=("", 10, "bold"), height=2, padx=10,
    ).pack(side="left", padx=15)

    def on_close():
        save_gui_config(snapshot_settings())  # 关闭窗口时记忆当前填写
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()


# =============================================================================
# 第九部分：命令行参数与程序入口
# =============================================================================

def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="县级土壤类型图多布局批量出图工具",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--mode", default="generate", choices=["generate", "export"],
                   help="generate=生成布局工程；export=批量导出 JPG")
    p.add_argument("--county", default="",
                   help="县/市名称（如：xx县），留空则从数据路径自动识别")
    p.add_argument("--mxd", default=str(CFG_FRAMES_BASE_DIR),
                   help="A3 图框底图 .mxd 路径")
    p.add_argument("--source-layer", default=str(CFG_SOIL_GDB_BASE_DIR),
                   help="土壤要素图层路径（.gdb/图层名）")
    p.add_argument("--color-table", default=str(CFG_COLOR_TABLE),
                   help="土壤类型配色表 CSV 路径")
    p.add_argument("--stats-table", default=str(CFG_STATS_BASE_DIR),
                   help="平差面积统计 Excel 路径（面积字段单位：平方米）")
    p.add_argument("--output-dir", default=str(CFG_OUTPUT_BASE_DIR),
                   help="成果输出目录（布局工程、JPG、.lyrx、配色对照表均输出至此）")
    p.add_argument("--aprx", default="",
                   help="第二步独立批量导出使用的 ArcGIS Pro 布局工程 .aprx；留空则使用输出目录中的自动化出图工作空间.aprx")
    p.add_argument("--soil-field", default=DEFAULT_SOIL_FIELD, help="土类字段名")
    p.add_argument("--subclass-field", default=DEFAULT_SUBCLASS_FIELD, help="亚类字段名")
    p.add_argument("--genus-field", default=DEFAULT_GENUS_FIELD,
                   help="土属字段名（留空则以亚类模式出图）")
    p.add_argument("--excludes", default=DEFAULT_EXCLUDES, help="排除土类，逗号分隔")
    p.add_argument("--area-unit", default=DEFAULT_AREA_UNIT, choices=["亩", "万亩"],
                   help="面积统计单位（配色对照表 Excel 中显示）")
    p.add_argument("--color-direction", default=DEFAULT_COLOR_DIRECTION,
                   choices=["dark_to_light", "light_to_dark"],
                   help="色带方向：dark_to_light=深→浅，light_to_dark=浅→深")
    p.add_argument("--dpi", type=int, default=DEFAULT_DPI, help="导出分辨率（DPI）")
    p.add_argument("--confirm-mismatch", action="store_true",
                   help="生成前发现统计表与图层类型不一致时弹窗确认")
    p.add_argument("--continue-on-mismatch", action="store_true",
                   help="生成前发现统计表与图层类型不一致时打印警告并继续")
    p.add_argument("--no-gui", action="store_true", help="跳过 GUI，直接执行命令行模式")
    return p.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    if not args.no_gui and len(sys.argv) == 1:
        run_gui()
        return 0
    outputs = export_with_arcgis_pro(args)
    for k, v in outputs.items():
        print(f"{k}: {v}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:
        print(f"错误：{exc}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)